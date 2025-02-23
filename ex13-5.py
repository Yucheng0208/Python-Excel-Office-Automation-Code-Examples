import openpyxl

file = "營業額.xlsx"
wb = openpyxl.load_workbook(file)
ws = wb.active

# 設定公式
ws["B10"] = "=SUM(B2:B9)"  # 總和
ws["B11"] = "=AVERAGE(B2:B9)"  # 平均

wb.save("營業額_計算.xlsx")
print("已儲存為 營業額_計算.xlsx")
